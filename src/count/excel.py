import datetime
import os
import openpyxl
import pandas as pd
import logging

import src.parameters

logger = logging.getLogger(__name__)

VERSION = src.parameters.VERSION


class Excel:

    def __init__(self, directory, counter):
        """Creates Excel file 'crypt_counts.xlsx' at given directory. If this
        excel file already exists, new saved data will be appended to a new
        sheet in the existing file. Sheet named by counter at creation.
        """
        self.time = str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        self.filepath = os.path.join(directory, "crypt_counts.xlsx")
        self.counter = counter
        self.empty = True
        # If Excel does not exist, create it
        if not os.path.exists(self.filepath):
            logger.info(f"Creating new Excel file at {self.filepath}.")
            self.__create_Excel__()
            return
        # Otherwise open it, open a new sheet
        wb = openpyxl.load_workbook(self.filepath)
        if self.counter in wb.sheetnames:
            logger.info("Current counter already exists, appending to that sheet.")
            self.empty = False
        else:
            logger.info("New counter - creating new sheet.")
            wb.create_sheet(self.counter)
        wb.save(self.filepath)
        # Write new counter info to version sheet.
        with pd.ExcelWriter(
            self.filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            startrow = writer.sheets["version"].max_row
            data = pd.DataFrame(
                {"Counter": [self.counter], "Timestamp": [self.time]}, index=[0]
            )
            data.to_excel(
                writer,
                sheet_name="version",
                index=False,
                header=False,
                startrow=startrow,
            )

    def __create_Excel__(self):
        """Creates an Excel file with the counter sheet and a version sheet."""
        # Open new workbook (xlsx file)
        wb = openpyxl.Workbook()
        # Change name of active sheet to current counter
        wb.active.title = self.counter
        # Create new sheet with name 'version'
        wb.create_sheet("version")
        # Save workbook to the same directory as the calibration film.
        wb.save(self.filepath)
        # Write header with version info and counter info to version sheet.
        with pd.ExcelWriter(
            self.filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            text = pd.DataFrame({f"Crypt GUI v{VERSION}": [" "]}, index=[0])
            text.to_excel(writer, sheet_name="version", index=False)
            data = pd.DataFrame(
                {"Counter": [self.counter], "Timestamp": [self.time]}, index=[0]
            )
            data.to_excel(
                writer,
                sheet_name="version",
                index=False,
                startrow=writer.sheets["version"].max_row,
            )

    def append(self, data):
        """Appends given data dictionary to xlsx file. If current excel is
        empty, creates a header first. If excel has a sheet with current
        counter, appends to that sheet.
        """
        with pd.ExcelWriter(
            self.filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            data = pd.DataFrame(data, index=[0])
            startrow = 0 if self.empty else writer.sheets[self.counter].max_row
            data.to_excel(
                writer,
                header=self.empty,
                startrow=startrow,
                index=False,
                sheet_name=self.counter,
            )
        self.empty = False  # From here on out, self.empty=False
